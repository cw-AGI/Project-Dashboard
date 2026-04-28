#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate VNPT 5G Central Coast workbook (5 provinces, 100–150 sites each,
new + expansion only) and optionally run export_dashboard.py.

Target dashboard: WN_Dashboard_v3.4_beautify.html
  cd "<repo>" && python3 -m http.server 8765
  open http://localhost:8765/WN_Dashboard_v3.4_beautify.html

Usage (from repo root):
  python scripts/generate_vnpt5g_central_workbook.py
  python scripts/generate_vnpt5g_central_workbook.py --no-export
"""

from __future__ import annotations

import argparse
import os
import random
import subprocess
import sys
from datetime import date, timedelta

PROJECT_ID = "vnpt5g_central"
REF_DATE = date(2026, 4, 27)  # reporting snapshot for status mix
PROJECT_START = date(2026, 1, 1)
PROJECT_END = date(2026, 4, 30)

# Five Central Coast provinces (names must match WN_Dashboard VN_PROVS / map)
PROVINCES = [
    {"name": "Đà Nẵng", "code": "DNG", "n_sites": 112, "lat0": 16.05, "lng0": 108.20},
    {"name": "Quảng Nam", "code": "QNM", "n_sites": 138, "lat0": 15.55, "lng0": 108.02},
    {"name": "Quảng Ngãi", "code": "QNI", "n_sites": 105, "lat0": 15.12, "lng0": 108.80},
    {"name": "Bình Định", "code": "BDI", "n_sites": 127, "lat0": 13.78, "lng0": 109.22},
    {"name": "Phú Yên", "code": "PHY", "n_sites": 119, "lat0": 13.10, "lng0": 109.32},
]

SITE_TYPES = [
    "5G NR n78 Macro",
    "5G NR n78 Q-Cell",
    "5G NR n2 Rural",
    "5G NR n78 Rooftop",
    "5G Expansion n78",
    "5G Expansion n2",
]


def _parse_iso(s: str) -> date:
    y, m, d = map(int, s.split("-"))
    return date(y, m, d)


def _status_for(ms: dict, ref: date) -> str:
    def ok(k: str) -> bool:
        v = ms.get(k)
        if not v:
            return False
        return _parse_iso(v) <= ref

    if ok("ms_PAT"):
        return "PAT Complete"
    if ok("ms_BTS_On_Air"):
        return "On Air Complete"
    if ok("ms_BTS_MOS"):
        return "Installation Complete"
    if ok("ms_RFI"):
        return "RFI Complete"
    return "RFI Complete"


def _milestone_row(tss_off: int) -> dict:
    """Fixed offsets from TSS (days); chain fits in Jan–Apr window."""
    base = PROJECT_START + timedelta(days=tss_off)
    d_tss = base
    d_tssr = base + timedelta(days=6)
    d_rfi = d_tssr + timedelta(days=8)
    d_mos = d_rfi + timedelta(days=16)
    d_on = d_mos + timedelta(days=16)
    d_pat = d_on + timedelta(days=12)
    return {
        "ms_TSS": d_tss.isoformat(),
        "ms_TSSR": d_tssr.isoformat(),
        "ms_RFI": d_rfi.isoformat(),
        "ms_BTS_MOS": d_mos.isoformat(),
        "ms_BTS_On_Air": d_on.isoformat(),
        "ms_PAT": d_pat.isoformat(),
    }


def _week_sundays() -> list[date]:
    """Sunday week-end dates from first Sunday on/after project start through end."""
    d = PROJECT_START
    while d.weekday() != 6:
        d += timedelta(days=1)
    out = []
    while d <= PROJECT_END + timedelta(days=7):
        out.append(d)
        d += timedelta(days=7)
    return out


def _iso_week_label(d: date) -> str:
    return f"W{d.isocalendar()[1]:02d}"


def _count_by_week(sites: list[dict], week_end: date, key: str) -> int:
    n = 0
    for s in sites:
        v = s.get(key)
        if not v:
            continue
        if _parse_iso(v) <= week_end:
            n += 1
    return n


def build_sites(rng: random.Random) -> list[dict]:
    sites = []
    seq = 0
    for prov in PROVINCES:
        n = prov["n_sites"]
        for i in range(1, n + 1):
            seq += 1
            tss_max = 52
            tss_off = int(rng.random() * (tss_max + 1))
            if rng.random() < 0.12:
                tss_off = min(tss_off + rng.randint(8, 22), 58)
            ms = _milestone_row(tss_off)
            if rng.random() < 0.04:
                del ms["ms_PAT"]
            elif rng.random() < 0.02:
                ms["ms_PAT"] = (PROJECT_END - timedelta(days=rng.randint(0, 2))).isoformat()

            bt = "new" if rng.random() < 0.62 else "expansion"
            lat = prov["lat0"] + rng.uniform(-0.28, 0.28)
            lng = prov["lng0"] + rng.uniform(-0.32, 0.32)
            site_id = f"{prov['code']}-{i:04d}"
            ward = f"Ward {((i * 17 + seq) % 40) + 1:02d}"
            sites.append(
                {
                    "site_id": site_id,
                    "site_name": f"{prov['code']} {ward} — {SITE_TYPES[rng.randrange(len(SITE_TYPES))]}",
                    "province": prov["name"],
                    "project": PROJECT_ID,
                    "site_type": SITE_TYPES[rng.randrange(len(SITE_TYPES))],
                    "build_type": bt,
                    "lat": round(lat, 6),
                    "lng": round(lng, 6),
                    **ms,
                    "status": "",
                }
            )
        for s in sites[-n:]:
            s["status"] = _status_for(s, REF_DATE)
    return sites


def build_rolling_rows(sites: list[dict]) -> list[dict]:
    rows = []
    total = len(sites)
    for we in _week_sundays():
        row = {
            "week": _iso_week_label(we),
            "week_end_date": we.isoformat(),
            "tss_plan": min(total, _count_by_week(sites, we, "ms_TSS")),
            "tssr_plan": min(total, _count_by_week(sites, we, "ms_TSSR")),
            "rfi_plan": min(total, _count_by_week(sites, we, "ms_RFI")),
            "install_plan": min(total, _count_by_week(sites, we, "ms_BTS_MOS")),
            "onair_plan": min(total, _count_by_week(sites, we, "ms_BTS_On_Air")),
            "pat_plan": min(total, _count_by_week(sites, we, "ms_PAT")),
            "pac_plan": 0,
            "ssv_plan": 0,
        }
        rows.append(row)
    return rows


def build_issues_weekly(rng: random.Random, rolling_rows: list[dict]) -> list[dict]:
    """Weekly raised/closed; export uses sum(raised), sum(closed), and last row's open."""
    weekly = []
    open_ = rng.randint(5, 9)
    for i, row in enumerate(rolling_rows):
        raised = rng.randint(2, 6)
        closed = rng.randint(1, max(1, raised - 1))
        if i > len(rolling_rows) * 0.55:
            closed = min(raised, closed + rng.randint(0, 2))
        open_ = max(0, min(22, open_ + raised - closed))
        weekly.append(
            {
                "week": row["week"],
                "raised": raised,
                "closed": closed,
                "open": open_,
            }
        )
    if weekly:
        weekly[-1]["open"] = min(weekly[-1]["open"], rng.randint(8, 14))
    return weekly


def build_resources_weekly(rng: random.Random, rolling_rows: list[dict]) -> list[dict]:
    weekly = []
    peak = 38
    n_weeks = len(rolling_rows)
    for w, rp in enumerate(rolling_rows):
        t = w / max(n_weeks - 1, 1)
        base = int(12 + (peak - 12) * (1 - abs(2 * t - 1) ** 0.9))
        install = max(8, base + rng.randint(-3, 4))
        ssv = max(1, install // 6 + rng.randint(0, 2))
        onair = max(1, install // 5 + rng.randint(0, 2))
        mw = max(1, install // 7)
        power = max(1, install // 8)
        optim = max(0, install // 10)
        total = install + ssv + onair + mw + power + optim
        weekly.append(
            {
                "week": rp["week"],
                "project": PROJECT_ID,
                "install": install,
                "ssv": ssv,
                "onair": onair,
                "mw": mw,
                "power": power,
                "optim": optim,
                "total": total,
                "total_plan": total + rng.randint(1, 4),
            }
        )
    return weekly


def write_workbook(path: str, sites: list[dict], rolling: list[dict]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError:
        print("ERROR: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    rng = random.Random(42)
    issues_w = build_issues_weekly(rng, rolling)
    res_w = build_resources_weekly(rng, rolling)

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws_pl = wb.create_sheet("PROJECT_LIST")
    ws_pl.append(["id", "name", "active", "color", "region"])
    ws_pl.append(
        [
            PROJECT_ID,
            "VNPT 5G — Central Coast (New + Expansion)",
            True,
            "#0f766e",
            "Miền Trung",
        ]
    )

    ws_cfg = wb.create_sheet("CONFIG")
    kv = [
        ("project_name", "VNPT 5G — Central Coast Rollout"),
        ("client", "VNPT"),
        ("start_date", PROJECT_START.isoformat()),
        ("end_date", PROJECT_END.isoformat()),
        ("required_complete_date", PROJECT_END.isoformat()),
        ("region", "Miền Trung — 5 tỉnh"),
        ("pm", "Nguyễn Văn Minh (PMO)"),
        ("weekly_target", 42),
        ("po_no", "PO-VNPT-5G-CM-2026-0142"),
        ("contract_value", "Khung hợp đồng triển khai 5G (simulated)"),
        ("vendor", "Nokia / Ericsson (đội triển khai liên danh)"),
        ("target_onair", len(sites)),
    ]
    for k, v in kv:
        ws_cfg.append([k, v])

    ws_data = wb.create_sheet("DATA")
    headers = [
        "site_id",
        "site_name",
        "province",
        "status",
        "project",
        "site_type",
        "build_type",
        "lat",
        "lng",
        "ms_TSS",
        "ms_TSSR",
        "ms_RFI",
        "ms_BTS_MOS",
        "ms_BTS_On_Air",
        "ms_PAT",
    ]
    ws_data.append(headers)
    for s in sites:
        ws_data.append([s.get(h) for h in headers])

    ws_iss = wb.create_sheet("ISSUES")
    ws_iss.append(["week", "raised", "closed", "open"])
    for row in issues_w:
        ws_iss.append([row["week"], row["raised"], row["closed"], row["open"]])

    ws_res = wb.create_sheet("RESOURCES")
    ws_res.append(
        ["week", "project", "install", "ssv", "onair", "mw", "power", "optim", "total", "total_plan"]
    )
    for row in res_w:
        ws_res.append(
            [
                row["week"],
                row["project"],
                row["install"],
                row["ssv"],
                row["onair"],
                row["mw"],
                row["power"],
                row["optim"],
                row["total"],
                row["total_plan"],
            ]
        )

    ws_risk = wb.create_sheet("RISKS")
    ws_risk.append(
        [
            "risk_name",
            "priority",
            "description",
            "owner",
            "resolve_date",
            "current_status",
            "next_steps",
            "project",
        ]
    )
    risks = [
        (
            "Spectrum Licensing and Local Permits",
            "Medium",
            "Some expansion sites in Quang Nam are delayed by local permits for auxiliary pole construction.",
            "Tran Hoai Nam",
            "2026-05-10",
            "The 12-site list is under review; 6 sites have received supplemental approvals.",
            "Coordinate with the provincial telecom authority and prioritize fast-track escalation in Week 18.",
            PROJECT_ID,
        ),
        (
            "Q1 RRH Equipment Logistics",
            "High",
            "The n78 RRH shipment missed port schedule windows, impacting the March installation sequence.",
            "Le Thu Ha",
            "2026-04-28",
            "Partial units were reallocated from the Singapore warehouse; around 70% of demand is now in the Central warehouse.",
            "Confirm air-freight for the final 40 RRHs and update the Week 17 MOS plan.",
            PROJECT_ID,
        ),
        (
            "Power and Shelter Readiness (Expansion)",
            "Medium",
            "Several expansion sites require power-capacity upgrades before On Air.",
            "Pham Quoc Anh",
            "2026-05-05",
            "Power survey is 89% complete; 11 sites are waiting for the electrical cabinet fabrication contract.",
            "Accelerate second-batch PDU procurement and lock installation slots for Weeks 16-17.",
            PROJECT_ID,
        ),
    ]
    for r in risks:
        ws_risk.append(list(r))

    sheet_rp_name = f"{PROJECT_ID.upper()}_ROLLING_PLAN"
    ws_rp = wb.create_sheet(sheet_rp_name)
    if rolling:
        ws_rp.append(list(rolling[0].keys()))
        for row in rolling:
            ws_rp.append(list(row.values()))

    wb.save(path)
    print(f"Wrote {path} ({len(sites)} sites)")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--output",
        default=os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "VNPT_5G_Central_Project.xlsx",
        ),
        help="Output .xlsx path",
    )
    ap.add_argument("--no-export", action="store_true", help="Skip export_dashboard.py")
    args = ap.parse_args()

    root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.chdir(root)

    rng = random.Random(202601)
    sites = build_sites(rng)
    rolling = build_rolling_rows(sites)

    out = os.path.abspath(args.output)
    write_workbook(out, sites, rolling)

    if not args.no_export:
        exp = os.path.join(root, "export_dashboard.py")
        r = subprocess.run(
            [sys.executable, exp, out, os.path.join(root, "projects")],
            cwd=root,
        )
        if r.returncode != 0:
            sys.exit(r.returncode)
        print("Export OK → projects/")


if __name__ == "__main__":
    main()
