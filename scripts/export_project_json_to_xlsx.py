#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Export a single project JSON folder into an editable dashboard workbook.

Usage:
  python scripts/export_project_json_to_xlsx.py --project-dir "projects/VNPT 5G PO2" --project-id "VNPT 5G PO2" --color "#0f766e" --output "VNPT_5G_PO2.xlsx"
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook


MILESTONE_KEYS = [
    "ms_TSS",
    "ms_TSSR",
    "ms_RFI",
    "ms_BTS_MOS",
    "ms_BTS_On_Air",
    "ms_PAT",
    "ms_PAC",
    "ms_SSV",
    "ms_MW_On_Air",
    "ms_Power_On_Air",
    "ms_TX_Readiness",
    "ms_BTS_HI",
    "ms_BTS_Integration",
    "ms_Dismantle",
    "ms_OSS_KPI",
    "ms_FAC",
]


def export_project(project_dir: Path, project_id: str, color: str, output_file: Path) -> None:
    data = json.loads((project_dir / "data.json").read_text(encoding="utf-8"))
    cfg = json.loads((project_dir / "config.json").read_text(encoding="utf-8"))

    wb = Workbook()
    wb.remove(wb.active)

    ws_pl = wb.create_sheet("PROJECT_LIST")
    ws_pl.append(["id", "name", "active", "color", "region"])
    ws_pl.append([project_id, cfg.get("project_name", project_id), True, color, cfg.get("region", "")])

    ws_cfg = wb.create_sheet("CONFIG")
    kv_order = [
        "project_name",
        "client",
        "start_date",
        "end_date",
        "required_complete_date",
        "region",
        "pm",
        "weekly_target",
        "po_no",
        "contract_value",
        "vendor",
        "target_onair",
    ]
    for key in kv_order:
        if key in cfg:
            ws_cfg.append([key, cfg.get(key)])

    ws_data = wb.create_sheet("DATA")
    data_headers = [
        "site_id",
        "site_name",
        "province",
        "status",
        "project",
        "site_type",
        "build_type",
        "lat",
        "lng",
    ] + MILESTONE_KEYS
    ws_data.append(data_headers)
    for site in data.get("sites", []):
        ws_data.append([site.get(h) for h in data_headers])

    ws_issues = wb.create_sheet("ISSUES")
    ws_issues.append(["week", "raised", "closed", "open"])
    for row in data.get("issues", {}).get("weekly", []):
        ws_issues.append([row.get("week"), row.get("raised"), row.get("closed"), row.get("open")])

    ws_res = wb.create_sheet("RESOURCES")
    res_headers = ["week", "project", "install", "ssv", "onair", "mw", "power", "optim", "total", "total_plan"]
    ws_res.append(res_headers)
    for row in data.get("resources", {}).get("weekly", []):
        ws_res.append([row.get(h) for h in res_headers])

    ws_risk = wb.create_sheet("RISKS")
    risk_headers = ["risk_name", "priority", "description", "owner", "resolve_date", "current_status", "next_steps", "project"]
    ws_risk.append(risk_headers)
    for row in data.get("risks", []):
        ws_risk.append([row.get(h) for h in risk_headers])

    rolling_rows = data.get("rolling_plan", {}).get("weekly", [])
    rp_name = f"{project_id.upper()}_ROLLING_PLAN"
    ws_rp = wb.create_sheet(rp_name)
    if rolling_rows:
        rp_headers = list(rolling_rows[0].keys())
        ws_rp.append(rp_headers)
        for row in rolling_rows:
            ws_rp.append([row.get(h) for h in rp_headers])

    wb.save(output_file)
    print(f"Wrote {output_file}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--project-dir", required=True)
    parser.add_argument("--project-id", required=True)
    parser.add_argument("--color", default="#1d4ed8")
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    export_project(
        project_dir=Path(args.project_dir),
        project_id=args.project_id,
        color=args.color,
        output_file=Path(args.output),
    )


if __name__ == "__main__":
    main()

