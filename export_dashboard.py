#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VN Dashboard Export Script  v1.0
=================================
Reads one Excel workbook and exports JSON files for GitHub Pages hosting.

Sheet layout expected in the workbook
--------------------------------------
  DATA          - One row per site (site_id, site_name, province, status,
                  project, site_type, build_type, lat, lng, ms_*, ext fields)
  CONFIG        - Key-value project info + FIELDS / EQUIP sections
  PROJECT_LIST  - Multi-project index (id, name, active, color, …)
  ISSUES        - Weekly issue tracking (week, raised, closed, open)
  RESOURCES     - Weekly resource tracking (week, install, mw, power, optim, total)

Output structure
----------------
  projects/
    index.json            ← merged from PROJECT_LIST, upserted each run
    <project_id>/
      data.json           ← sites + issues + resources
      config.json         ← project metadata + field definitions

Usage
-----
  python export_dashboard.py                          # auto-detect .xlsx
  python export_dashboard.py MyProject.xlsx           # explicit file
  python export_dashboard.py MyProject.xlsx ./output  # custom output dir
"""

import json
import glob
import os
import sys
from datetime import datetime, date


# ─────────────────────────────────────────────────────────────
# MILESTONE DEFINITIONS  (fixed, not read from Excel)
# ─────────────────────────────────────────────────────────────
MILESTONES = [
    {"key": "ms_TSS",             "label": "TSS",            "category": "planning"},
    {"key": "ms_TSSR",            "label": "TSSR",           "category": "planning"},
    {"key": "ms_RFI",             "label": "RFI",            "category": "delivery"},
    {"key": "ms_BTS_MOS",         "label": "MOS",            "category": "delivery"},
    {"key": "ms_BTS_On_Air",      "label": "On Air",         "category": "delivery"},
    {"key": "ms_PAT",             "label": "PAT",            "category": "acceptance"},
    {"key": "ms_PAC",             "label": "PAC",            "category": "acceptance"},
    {"key": "ms_SSV",             "label": "SSV",            "category": "acceptance"},
    {"key": "ms_MW_On_Air",       "label": "MW On Air",      "scope": ["swap", "new"], "category": "support"},
    {"key": "ms_Power_On_Air",    "label": "Power On Air",   "category": "support"},
    {"key": "ms_TX_Readiness",    "label": "TX Readiness",   "category": "support"},
    {"key": "ms_BTS_HI",          "label": "BTS HI",         "category": "support"},
    {"key": "ms_BTS_Integration", "label": "Integration",    "category": "support"},
    {"key": "ms_Dismantle",       "label": "Dismantle",      "scope": ["swap"], "category": "support"},
    {"key": "ms_OSS_KPI",         "label": "OSS KPI",        "category": "acceptance"},
    {"key": "ms_FAC",             "label": "FAC",            "category": "acceptance"},
]

BUILD_TYPES = [
    {"key": "new",       "label": "New Sites"},
    {"key": "swap",      "label": "SWAP Sites"},
    {"key": "expansion", "label": "Expansion Sites"},
]

STATUS_COLORS = {
    "RFI Complete":          "#FFC107",
    "Installation Complete": "#42A5F5",
    "On Air Complete":       "#66BB6A",
    "PAT Complete":          "#AB47BC",
}

DEFAULT_OVERVIEW_MILESTONES = [
    "ms_TSSR",
    "ms_RFI",
    "ms_BTS_MOS",
    "ms_BTS_On_Air",
]

DEFAULT_HIDDEN_MILESTONES = [
    "ms_MW_On_Air",
    "ms_Power_On_Air",
    "ms_TX_Readiness",
    "ms_BTS_HI",
    "ms_BTS_Integration",
    "ms_Dismantle",
    "ms_OSS_KPI",
    "ms_FAC",
]


# ─────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────
def _serialize(obj):
    """JSON default serializer for date/datetime objects."""
    if isinstance(obj, (datetime, date)):
        return obj.strftime("%Y-%m-%d")
    return str(obj)


def write_json(path, data):
    dir_ = os.path.dirname(path)
    if dir_:
        os.makedirs(dir_, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2, default=_serialize)
    print(f"    ✓  {path}")


def cell_val(cell):
    """Return a cell's value, converting date/datetime → 'YYYY-MM-DD' string."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    # Strip accidental float tails for pure integers stored as float
    if isinstance(v, float) and v == int(v):
        return int(v)
    return v


def sheet_to_dicts(ws, min_row=2):
    """Convert a sheet with a header row into a list of dicts."""
    headers = [cell_val(c) for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=min_row):
        if not row[0].value:          # skip blank rows
            continue
        entry = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                entry[h] = cell_val(row[i])
        rows.append(entry)
    return rows


# ─────────────────────────────────────────────────────────────
# SHEET READERS
# ─────────────────────────────────────────────────────────────
def read_project_list(ws):
    """Return list of project dicts from PROJECT_LIST sheet."""
    rows = sheet_to_dicts(ws)
    for p in rows:
        # Normalise the 'active' column (TRUE / FALSE / 1 / 0)
        raw = str(p.get("active", "true")).strip().upper()
        p["active"] = raw in ("TRUE", "1", "YES")
    return rows


def _normalise_bool(value, default=True):
    if value is None:
        return default
    return str(value).strip().upper() not in ("FALSE", "0", "NO", "N")


def read_config(ws):
    """
    Parse CONFIG sheet into a dict.

    Structure expected:
      Rows 1–N   : key | value   (plain key-value pairs)
      Row ?      : FIELDS_START  (marker; following rows = field definitions)
                   key | label | type | visible
      Row ?      : EQUIP_START   (marker; following rows = equipment)
                   (ignored col) | type | model
    """
    config = {}
    fields = []
    equipment = []
    mode = "kv"   # 'kv' | 'fields_hdr' | 'fields' | 'equip'

    for row in ws.iter_rows():
        a = cell_val(row[0]) if len(row) > 0 else None
        b = cell_val(row[1]) if len(row) > 1 else None
        c = cell_val(row[2]) if len(row) > 2 else None
        d = cell_val(row[3]) if len(row) > 3 else None

        # Section markers
        if a == "FIELDS_START":
            mode = "fields_hdr"
            continue
        if a == "EQUIP_START":
            mode = "equip"
            continue

        if mode == "kv":
            if a and b is not None:
                config[str(a).strip()] = b

        elif mode == "fields_hdr":
            # First row after FIELDS_START is the column header row; skip it
            mode = "fields"

        elif mode == "fields":
            if a:   # a = key
                field = {
                    "key":   str(a).strip(),
                    "label": str(b).strip() if b else str(a).strip(),
                    "type":  str(c).strip() if c else "text",
                }
                if d is not None:
                    field["visible"] = str(d).strip().upper() != "FALSE"
                fields.append(field)

        elif mode == "equip":
            if b:   # b = type
                equipment.append({
                    "type":  str(b).strip(),
                    "model": str(c).strip() if c else "",
                })

    # Attach fixed definitions
    config["milestones"]   = MILESTONES
    config["build_types"]  = BUILD_TYPES
    config["status_colors"] = STATUS_COLORS
    config["overview_milestones"] = DEFAULT_OVERVIEW_MILESTONES
    config["hidden_milestones"] = DEFAULT_HIDDEN_MILESTONES
    if fields:
        config["fields"] = fields
    if equipment:
        config["equipment"] = equipment

    return config


def read_data(ws, project_id):
    """Return list of site dicts filtered to project_id."""
    all_rows = sheet_to_dicts(ws)
    sites = [r for r in all_rows if str(r.get("project", "")).strip() == project_id]
    # Normalise build_type and site_type to lowercase for consistency
    for s in sites:
        if s.get("build_type"):
            s["build_type"] = str(s["build_type"]).strip().lower()
        if s.get("site_type"):
            s["site_type"] = str(s["site_type"]).strip()
    return sites


def read_issues(ws):
    """
    Return issues summary + weekly list.
    Totals: raised = sum of all raised column; closed = sum closed; open = last row's open.
    """
    if ws is None:
        return {"raised": 0, "closed": 0, "open": 0, "weekly": []}

    weekly = sheet_to_dicts(ws)
    if not weekly:
        return {"raised": 0, "closed": 0, "open": 0, "weekly": []}

    total_raised = sum(int(w.get("raised") or 0) for w in weekly)
    total_closed = sum(int(w.get("closed") or 0) for w in weekly)
    last_open    = int(weekly[-1].get("open") or 0)

    return {
        "raised": total_raised,
        "closed": total_closed,
        "open":   last_open,
        "weekly": weekly,
    }


def _filter_project_rows(rows, project_id):
    project_rows = []
    for row in rows:
        row_project = str(row.get("project", "")).strip().lower()
        if not row_project or row_project == str(project_id).strip().lower():
            project_rows.append(row)
    return project_rows


def read_resources(ws, project_id):
    """Return resources weekly list, optionally filtered by project."""
    if ws is None:
        return {"weekly": []}
    rows = sheet_to_dicts(ws)
    return {"weekly": _filter_project_rows(rows, project_id)}


def read_rolling_plan(wb, project_id):
    """Read project rolling plan from a dedicated sheet like VNPT_ROLLING_PLAN."""
    sheet_name = f"{str(project_id).strip().upper()}_ROLLING_PLAN"
    if sheet_name not in wb.sheetnames:
        return {"sheet": sheet_name, "weekly": []}

    rows = sheet_to_dicts(wb[sheet_name])
    normalised = []
    for row in rows:
        item = {}
        for key, value in row.items():
            item[str(key).strip()] = value
        normalised.append(item)
    return {"sheet": sheet_name, "weekly": normalised}


def read_risks(ws, project_id):
    """Read structured risks list and keep rows matching the target project when provided."""
    if ws is None:
        return []

    rows = sheet_to_dicts(ws)
    filtered = _filter_project_rows(rows, project_id)

    def priority_rank(value):
        v = str(value or "").strip().lower()
        return {"critical": 0, "high": 1, "medium": 2, "low": 3}.get(v, 9)

    filtered.sort(
        key=lambda row: (
            priority_rank(row.get("priority")),
            str(row.get("resolve_date") or "9999-99-99"),
            str(row.get("risk_name") or row.get("title") or ""),
        )
    )
    return filtered


# ─────────────────────────────────────────────────────────────
# INDEX.JSON UPSERT
# ─────────────────────────────────────────────────────────────
def upsert_index(index_path, entry):
    """Load existing index.json, update/insert entry, write back."""
    existing = []
    if os.path.exists(index_path):
        try:
            with open(index_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass

    updated = False
    for i, p in enumerate(existing):
        if p.get("id") == entry["id"]:
            existing[i] = entry
            updated = True
            break
    if not updated:
        existing.append(entry)

    write_json(index_path, existing)


# ─────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────
def main():
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl is not installed.")
        print("       Run:  pip install openpyxl --break-system-packages")
        sys.exit(1)

    # ── Resolve paths ───────────────────────────────────────
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        candidates = glob.glob("*.xlsx")
        if not candidates:
            print("ERROR: No .xlsx file found in current directory.")
            print("Usage: python export_dashboard.py [excel_path] [output_dir]")
            sys.exit(1)
        excel_path = candidates[0]

    output_dir = sys.argv[2] if len(sys.argv) > 2 else "projects"

    print()
    print("=" * 50)
    print("  VN Dashboard Export")
    print("=" * 50)
    print(f"  Source : {excel_path}")
    print(f"  Output : {output_dir}/")
    print(f"  Time   : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()

    # ── Load workbook ───────────────────────────────────────
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        print(f"ERROR loading workbook: {e}")
        sys.exit(1)

    # Required sheets
    for name in ("DATA", "CONFIG", "PROJECT_LIST"):
        if name not in wb.sheetnames:
            print(f"ERROR: Required sheet '{name}' not found.")
            print(f"       Available sheets: {wb.sheetnames}")
            sys.exit(1)

    ws_data      = wb["DATA"]
    ws_config    = wb["CONFIG"]
    ws_proj_list = wb["PROJECT_LIST"]
    ws_issues    = wb["ISSUES"]    if "ISSUES"    in wb.sheetnames else None
    ws_resources = wb["RESOURCES"] if "RESOURCES" in wb.sheetnames else None
    ws_risks     = wb["RISKS"]     if "RISKS"     in wb.sheetnames else None

    if ws_issues is None:
        print("  ⚠  Sheet 'ISSUES' not found – issues data will be empty.")
    if ws_resources is None:
        print("  ⚠  Sheet 'RESOURCES' not found – resources data will be empty.")
    if ws_risks is None:
        print("  ⚠  Sheet 'RISKS' not found – top risks data will be empty.")

    # ── Read project list ────────────────────────────────────
    projects = read_project_list(ws_proj_list)
    active   = [p for p in projects if p.get("active", True)]
    print(f"  Projects found : {[p['id'] for p in projects]}")
    print(f"  Active         : {[p['id'] for p in active]}")
    print()

    # ── Export each active project ───────────────────────────
    for proj in active:
        pid = str(proj["id"]).strip()
        print(f"  ── {pid} ──────────────────────────────")

        proj_dir = os.path.join(output_dir, pid)

        # Config
        config = read_config(ws_config)
        config["project"] = pid   # ensure project field is set
        config["project_name"] = proj.get("name", config.get("project_name", pid))
        if proj.get("region"):
            config["region"] = proj.get("region")

        # Sites
        sites = read_data(ws_data, pid)
        print(f"     Sites    : {len(sites)}")

        # Issues, Resources, Rolling Plan, Risks
        issues    = read_issues(ws_issues)
        resources = read_resources(ws_resources, pid)
        rolling_plan = read_rolling_plan(wb, pid)
        risks = read_risks(ws_risks, pid)
        print(f"     Issues   : raised={issues['raised']}  closed={issues['closed']}  open={issues['open']}")
        print(f"     Resources: {len(resources['weekly'])} weeks")
        print(f"     Rolling  : {len(rolling_plan['weekly'])} weeks ({rolling_plan['sheet']})")
        print(f"     Risks    : {len(risks)} items")

        # data.json
        data_export = {
            "generated": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "project":   pid,
            "sites":     sites,
            "issues":    issues,
            "resources": resources,
            "rolling_plan": rolling_plan,
            "risks": risks,
        }
        write_json(os.path.join(proj_dir, "data.json"),   data_export)
        write_json(os.path.join(proj_dir, "config.json"), config)

        # index.json (upsert)
        index_entry = {
            "id":      pid,
            "name":    proj.get("name", pid),
            "path":    f"projects/{pid}/",
            "color":   proj.get("color", "#1d4ed8"),
            "active":  True,
            "updated": datetime.now().strftime("%Y-%m-%d"),
        }
        upsert_index(os.path.join(output_dir, "index.json"), index_entry)
        print()

    print("=" * 50)
    print("  ✅  Export complete!")
    print("  → git add projects/ && git push")
    print("=" * 50)
    print()


if __name__ == "__main__":
    main()
